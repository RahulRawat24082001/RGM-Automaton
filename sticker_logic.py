"""
Core logic for generating Job Card sticker numbers.

Numbering rule:
    sticker number = <Job Card No> + <last 2 digits of F.Year> + <6-digit running sequence>
    e.g. Job Card No 4, F.Year "26-27", sequence 1  ->  426000001

Layout rule (Stickers sheet):
    - Numbers are filled DOWN a column, ROWS_PER_COL at a time, then continue in the next column.
    - Every new Job Card ALWAYS starts at row 1 of a fresh column
      (even if the previous job card's last column wasn't full).
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

ROWS_PER_COL = 47           # matches the original template (47 stickers per column)
DETAILS_SHEET_NAME = "Sticker Details"
STICKERS_SHEET_NAME = "Stickers"

COL_JOB_CARD = 3   # "Job Card No"
COL_F_YEAR = 4     # "F. Year"
COL_LOT_QTY = 5    # "Lot Qty"


def _year_suffix(f_year_value) -> str:
    """Extract the 2-digit prefix year from a value like '26-27' -> '26'."""
    text = str(f_year_value).strip()
    first_part = text.split("-")[0].strip()
    digits = "".join(ch for ch in first_part if ch.isdigit())
    return digits[-2:].zfill(2) if digits else "00"


def read_job_cards(ws_details):
    """Read (job_card_no, year_suffix, qty, row_index) for every valid data row."""
    jobs = []
    for r in range(2, ws_details.max_row + 1):
        job = ws_details.cell(r, COL_JOB_CARD).value
        f_year = ws_details.cell(r, COL_F_YEAR).value
        qty = ws_details.cell(r, COL_LOT_QTY).value
        if job is None or qty is None:
            continue
        try:
            qty = int(qty)
        except (ValueError, TypeError):
            continue
        if qty <= 0:
            continue
        jobs.append((job, _year_suffix(f_year), qty, r))
    return jobs


def build_stickers_sheet(wb, jobs):
    """Delete any existing Stickers sheet and rebuild it from the job list."""
    if STICKERS_SHEET_NAME in wb.sheetnames:
        del wb[STICKERS_SHEET_NAME]
    ws = wb.create_sheet(STICKERS_SHEET_NAME)

    bold_font = Font(name="Calibri", size=11, bold=True)
    center = Alignment(horizontal="center", wrap_text=True)

    col = 1
    row = 1
    max_col_used = 1

    for job, yr_suffix, qty, _src_row in jobs:
        prefix = f"{job}{yr_suffix}"
        # Sequence width = number of digits in the lot qty, plus 2
        # (e.g. qty 4000 -> 4 digits -> width 6 -> 000001..004000;
        #       qty 500  -> 3 digits -> width 5 -> 00001..00500)
        seq_width = len(str(qty)) + 2
        # Each job card always starts at a fresh column, row 1
        if row != 1:
            col += 1
            row = 1

        for seq in range(1, qty + 1):
            sticker_no = int(f"{prefix}{seq:0{seq_width}d}")
            cell = ws.cell(row=row, column=col, value=sticker_no)
            cell.font = bold_font
            cell.alignment = center
            max_col_used = max(max_col_used, col)

            row += 1
            if row > ROWS_PER_COL:
                row = 1
                col += 1

    for c in range(1, max_col_used + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10

    return ws


def process_workbook(input_path_or_buffer, output_path):
    """Load workbook, rebuild Stickers sheet, save to output_path. Returns summary list."""
    wb = load_workbook(input_path_or_buffer)

    if DETAILS_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{DETAILS_SHEET_NAME}' not found in the uploaded file.")

    ws_details = wb[DETAILS_SHEET_NAME]
    jobs = read_job_cards(ws_details)

    if not jobs:
        raise ValueError("No valid Job Card rows (with Job Card No and Lot Qty) were found.")

    build_stickers_sheet(wb, jobs)
    wb.save(output_path)

    summary = []
    for job, yr_suffix, qty, src_row in jobs:
        prefix = f"{job}{yr_suffix}"
        first_no = f"{prefix}{1:06d}"
        last_no = f"{prefix}{qty:06d}"
        summary.append(
            {
                "Job Card No": job,
                "F. Year": ws_details.cell(src_row, COL_F_YEAR).value,
                "Lot Qty": qty,
                "First Sticker": first_no,
                "Last Sticker": last_no,
            }
        )
    return summary
