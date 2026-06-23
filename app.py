import io
import tempfile
import os

import streamlit as st
from openpyxl import load_workbook

from sticker_logic import process_workbook, DETAILS_SHEET_NAME, STICKERS_SHEET_NAME

st.set_page_config(page_title="Job Card Sticker Generator", layout="centered")

st.title("🏷️ Job Card Sticker Number Generator")
st.write(
    "Upload your Excel sheet (with a **Sticker Details** sheet listing Job Card No, "
    "F. Year and Lot Qty). This tool will regenerate the **Stickers** sheet with the "
    "correct sticker numbers for every job card."
)

with st.expander("How the numbering works"):
    st.markdown(
        """
- **Sticker number** = `Job Card No` + last 2 digits of `F. Year` + a running sequence number.
- The sequence is zero-padded to *(number of digits in Lot Qty) + 2*.
  e.g. Lot Qty `4000` (4 digits) → 6-digit sequence → `000001` … `004000`.
  Lot Qty `500` (3 digits) → 5-digit sequence → `00001` … `00500`.
- Numbers fill **down** a column, 47 per column, then continue in the next column.
- **Every new Job Card always starts at row 1 of a brand-new column.**
        """
    )

uploaded_file = st.file_uploader("Upload Excel file (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    try:
        # Quick preview of detected job cards
        preview_wb = load_workbook(uploaded_file, data_only=True)
        if DETAILS_SHEET_NAME not in preview_wb.sheetnames:
            st.error(f"Could not find a sheet named '{DETAILS_SHEET_NAME}' in this file.")
        else:
            uploaded_file.seek(0)

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_in:
                tmp_in.write(uploaded_file.getvalue())
                input_path = tmp_in.name

            output_path = input_path.replace(".xlsx", "_output.xlsx")

            with st.spinner("Generating sticker numbers..."):
                summary = process_workbook(input_path, output_path)

            st.success(f"Done! Generated stickers for {len(summary)} job card(s).")

            st.subheader("Summary")
            st.dataframe(summary, use_container_width=True)

            with open(output_path, "rb") as f:
                output_bytes = f.read()

            st.download_button(
                label="⬇️ Download updated Excel file",
                data=output_bytes,
                file_name="RGM_Sticker_Updated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            os.remove(input_path)
            os.remove(output_path)

    except Exception as e:
        st.error(f"Something went wrong: {e}")
else:
    st.info("Waiting for a file to be uploaded.")
